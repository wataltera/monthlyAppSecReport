from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file
import sqlite3
from datetime import datetime
import os
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import io

app = Flask(__name__)
app.secret_key = 'dev-secret-key-change-in-production'

DB_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'monthlyReport.db')

def get_db_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/artifacts')
def artifacts():
    conn = get_db_connection()
    
    # Clear filters if requested
    if request.args.get('clear'):
        session.pop('artifact_filters', None)
        return redirect(url_for('artifacts'))
    
    # Get filter parameters from query string or session
    if request.args:
        # Store filters in session when provided via query string
        session['artifact_filters'] = {
            'business_unit': request.args.get('business_unit', '').strip(),
            'altera_product': request.args.get('altera_product', '').strip(),
            'rapid7_app': request.args.get('rapid7_app', '').strip(),
            'checkmarx_product': request.args.get('checkmarx_product', '').strip(),
            'mend_product': request.args.get('mend_product', '').strip(),
            'mend_project': request.args.get('mend_project', '').strip()
        }
    
    # Use session filters or empty defaults
    filters = session.get('artifact_filters', {})
    business_unit = filters.get('business_unit', '')
    altera_product = filters.get('altera_product', '')
    rapid7_app = filters.get('rapid7_app', '')
    checkmarx_product = filters.get('checkmarx_product', '')
    mend_product = filters.get('mend_product', '')
    mend_project = filters.get('mend_project', '')
    
    # Build query with filters
    query = 'SELECT * FROM Artifacts WHERE Deleted = 0'
    params = []
    
    if business_unit:
        query += ' AND BusinessUnit = ?'
        params.append(business_unit)
    
    if altera_product:
        query += ' AND AlteraProduct LIKE ?'
        params.append(f'%{altera_product}%')
    
    if rapid7_app:
        query += ' AND Rapid7App LIKE ?'
        params.append(f'%{rapid7_app}%')
    
    if checkmarx_product:
        query += ' AND CheckmarxProduct LIKE ?'
        params.append(f'%{checkmarx_product}%')
    
    if mend_product:
        query += ' AND MendProduct LIKE ?'
        params.append(f'%{mend_product}%')
    
    if mend_project:
        query += ' AND MendProject LIKE ?'
        params.append(f'%{mend_project}%')
    
    query += ' ORDER BY BusinessUnit, ID'
    
    artifacts = conn.execute(query, params).fetchall()
    conn.close()
    return render_template('artifacts.html', artifacts=artifacts, filters=filters)

@app.route('/artifacts/export')
def export_artifacts():
    conn = get_db_connection()
    
    # Use session filters
    filters = session.get('artifact_filters', {})
    business_unit = filters.get('business_unit', '')
    altera_product = filters.get('altera_product', '')
    rapid7_app = filters.get('rapid7_app', '')
    checkmarx_product = filters.get('checkmarx_product', '')
    mend_product = filters.get('mend_product', '')
    mend_project = filters.get('mend_project', '')
    
    # Build query with same filters
    query = 'SELECT * FROM Artifacts WHERE Deleted = 0'
    params = []
    
    if business_unit:
        query += ' AND BusinessUnit = ?'
        params.append(business_unit)
    
    if altera_product:
        query += ' AND AlteraProduct LIKE ?'
        params.append(f'%{altera_product}%')
    
    if rapid7_app:
        query += ' AND Rapid7App LIKE ?'
        params.append(f'%{rapid7_app}%')
    
    if checkmarx_product:
        query += ' AND CheckmarxProduct LIKE ?'
        params.append(f'%{checkmarx_product}%')
    
    if mend_product:
        query += ' AND MendProduct LIKE ?'
        params.append(f'%{mend_product}%')
    
    if mend_project:
        query += ' AND MendProject LIKE ?'
        params.append(f'%{mend_project}%')
    
    query += ' ORDER BY BusinessUnit, ID'
    
    artifacts = conn.execute(query, params).fetchall()
    conn.close()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    
    # Build sheet name from filters
    sheet_name_parts = []
    if business_unit:
        sheet_name_parts.append(business_unit)
    else:
        sheet_name_parts.append('AllBUs')
    
    if altera_product:
        sheet_name_parts.append(f'Altera={altera_product}')
    if rapid7_app:
        sheet_name_parts.append(f'R7={rapid7_app}')
    if checkmarx_product:
        sheet_name_parts.append(f'CX={checkmarx_product}')
    if mend_product:
        sheet_name_parts.append(f'MP={mend_product}')
    if mend_project:
        sheet_name_parts.append(f'MJ={mend_project}')
    
    sheet_name = '_'.join(sheet_name_parts)[:31]  # Excel sheet name limit
    ws.title = sheet_name
    
    # Headers
    headers = ['ID', 'BusinessUnit', 'AlteraProduct', 'Rapid7App', 'CheckmarxProduct', 'MendProduct', 
               'MendProject', 'Owner', 'SCAScans', 'SASTScans', 'DASTScans', 'RecentSCA', 
               'RecentSCAOK', 'RecentSAST', 'RecentSASTOK', 'RecentDAST', 'RecentDASTOK', 'RecentLOC']
    ws.append(headers)
    
    # Make header bold
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    # Add data
    for artifact in artifacts:
        ws.append([
            artifact['ID'], artifact['BusinessUnit'], artifact['AlteraProduct'], 
            artifact['Rapid7App'], artifact['CheckmarxProduct'], artifact['MendProduct'],
            artifact['MendProject'], artifact['Owner'], artifact['SCAScans'], 
            artifact['SASTScans'], artifact['DASTScans'], artifact['RecentSCA'],
            artifact['RecentSCAOK'], artifact['RecentSAST'], artifact['RecentSASTOK'],
            artifact['RecentDAST'], artifact['RecentDASTOK'], artifact['RecentLOC']
        ])
    
    # Freeze top row
    ws.freeze_panes = 'A2'
    
    # Add autofilter
    ws.auto_filter.ref = ws.dimensions
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'Artifacts_{sheet_name}_{timestamp}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/artifacts/new', methods=['GET', 'POST'])
def new_artifact():
    if request.method == 'POST':
        try:
            # Convert empty strings to None, but keep non-empty values
            def clean_value(val):
                val = val.strip() if val else ''
                return val if val else None
            
            conn = get_db_connection()
            conn.execute('''
                INSERT INTO Artifacts (
                    BusinessUnit, AlteraProduct, Rapid7App, CheckmarxProduct, MendProduct, MendProject, Owner,
                    SCAScans, SASTScans, DASTScans, RecentSCA, RecentSCAOK, RecentSAST, RecentSASTOK,
                    RecentDAST, RecentDASTOK, RecentLOC, Deleted
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                request.form['business_unit'].strip(),
                clean_value(request.form.get('altera_product', '')),
                clean_value(request.form.get('rapid7_app', '')),
                clean_value(request.form.get('checkmarx_product', '')),
                clean_value(request.form.get('mend_product', '')),
                clean_value(request.form.get('mend_project', '')),
                clean_value(request.form.get('owner', '')),
                int(request.form.get('sca_scans', 0)),
                int(request.form.get('sast_scans', 0)),
                int(request.form.get('dast_scans', 0)),
                clean_value(request.form.get('recent_sca', '')),
                int(request.form.get('recent_sca_ok', 0)),
                clean_value(request.form.get('recent_sast', '')),
                int(request.form.get('recent_sast_ok', 0)),
                clean_value(request.form.get('recent_dast', '')),
                int(request.form.get('recent_dast_ok', 0)),
                int(request.form.get('recent_loc', 0)),
                0
            ))
            conn.commit()
            conn.close()
            flash('Artifact created successfully!', 'success')
            return redirect(url_for('artifacts'))
        except sqlite3.IntegrityError as e:
            flash(f'Database error: {str(e)}', 'error')
        except Exception as e:
            flash(f'Error: {str(e)}', 'error')
    
    return render_template('artifact_form.html', artifact=None)

@app.route('/artifacts/<int:id>/edit', methods=['GET', 'POST'])
def edit_artifact(id):
    conn = get_db_connection()
    
    if request.method == 'POST':
        try:
            # Convert empty strings to None, but keep non-empty values
            def clean_value(val):
                val = val.strip() if val else ''
                return val if val else None
            
            conn.execute('''
                UPDATE Artifacts SET
                    BusinessUnit = ?, AlteraProduct = ?, Rapid7App = ?, CheckmarxProduct = ?, MendProduct = ?, 
                    MendProject = ?, Owner = ?, SCAScans = ?, SASTScans = ?, DASTScans = ?,
                    RecentSCA = ?, RecentSCAOK = ?, RecentSAST = ?, RecentSASTOK = ?,
                    RecentDAST = ?, RecentDASTOK = ?, RecentLOC = ?
                WHERE ID = ?
            ''', (
                request.form['business_unit'].strip(),
                clean_value(request.form.get('altera_product', '')),
                clean_value(request.form.get('rapid7_app', '')),
                clean_value(request.form.get('checkmarx_product', '')),
                clean_value(request.form.get('mend_product', '')),
                clean_value(request.form.get('mend_project', '')),
                clean_value(request.form.get('owner', '')),
                int(request.form.get('sca_scans', 0)),
                int(request.form.get('sast_scans', 0)),
                int(request.form.get('dast_scans', 0)),
                clean_value(request.form.get('recent_sca', '')),
                int(request.form.get('recent_sca_ok', 0)),
                clean_value(request.form.get('recent_sast', '')),
                int(request.form.get('recent_sast_ok', 0)),
                clean_value(request.form.get('recent_dast', '')),
                int(request.form.get('recent_dast_ok', 0)),
                int(request.form.get('recent_loc', 0)),
                id
            ))
            conn.commit()
            conn.close()
            flash('Artifact updated successfully!', 'success')
            return redirect(url_for('artifacts'))
        except sqlite3.IntegrityError as e:
            flash(f'Database error: {str(e)}', 'error')
        except Exception as e:
            flash(f'Error: {str(e)}', 'error')
    
    artifact = conn.execute('SELECT * FROM Artifacts WHERE ID = ?', (id,)).fetchone()
    conn.close()
    
    if artifact is None:
        flash('Artifact not found!', 'error')
        return redirect(url_for('artifacts'))
    
    return render_template('artifact_form.html', artifact=artifact)

@app.route('/artifacts/<int:id>/delete', methods=['POST'])
def delete_artifact(id):
    conn = get_db_connection()
    conn.execute('UPDATE Artifacts SET Deleted = 1 WHERE ID = ?', (id,))
    conn.commit()
    conn.close()
    flash('Artifact marked as deleted!', 'success')
    return redirect(url_for('artifacts'))

@app.route('/scans')
def scans():
    conn = get_db_connection()
    
    # Clear filters if requested
    if request.args.get('clear'):
        session.pop('scan_filters', None)
        return redirect(url_for('scans'))
    
    # Get filter parameters from query string or session
    if request.args:
        # Store filters in session when provided via query string
        session['scan_filters'] = {
            'business_unit': request.args.get('business_unit', '').strip(),
            'scan_tool': request.args.get('scan_tool', '').strip(),
            'scan_type': request.args.get('scan_type', '').strip(),
            'most_recent_only': request.args.get('most_recent_only') == '1'
        }
    
    # Use session filters or empty defaults
    filters = session.get('scan_filters', {})
    business_unit = filters.get('business_unit', '')
    scan_tool = filters.get('scan_tool', '')
    scan_type = filters.get('scan_type', '')
    most_recent_only = filters.get('most_recent_only', False)
    
    # Build query with filters
    if most_recent_only:
        # Get most recent scan for each artifact/tool/type combination
        query = '''
            SELECT s.*, a.BusinessUnit, a.Rapid7App, a.CheckmarxProduct, a.MendProduct, a.MendProject
            FROM Scans s
            JOIN Artifacts a ON s.ArtifactID = a.ID
            WHERE s.ID IN (
                SELECT s2.ID FROM Scans s2
                WHERE s2.ArtifactID = s.ArtifactID 
                  AND s2.ScanTool = s.ScanTool
                  AND s2.ScanType = s.ScanType
                  AND s2.ScanDateTime = (
                    SELECT MAX(s3.ScanDateTime) 
                    FROM Scans s3
                    WHERE s3.ArtifactID = s2.ArtifactID
                      AND s3.ScanTool = s2.ScanTool
                      AND s3.ScanType = s2.ScanType
                  )
            )
        '''
    else:
        query = '''
            SELECT s.*, a.BusinessUnit, a.Rapid7App, a.CheckmarxProduct, a.MendProduct, a.MendProject
            FROM Scans s
            JOIN Artifacts a ON s.ArtifactID = a.ID
            WHERE 1=1
        '''
    params = []
    
    if business_unit:
        query += ' AND a.BusinessUnit = ?'
        params.append(business_unit)
    
    if scan_tool:
        query += ' AND s.ScanTool LIKE ?'
        params.append(f'%{scan_tool}%')
    
    if scan_type:
        query += ' AND s.ScanType LIKE ?'
        params.append(f'%{scan_type}%')
    
    query += ' ORDER BY s.ScanDateTime DESC'
    
    scans = conn.execute(query, params).fetchall()
    conn.close()
    return render_template('scans.html', scans=scans, filters=filters)

@app.route('/scans/export')
def export_scans():
    conn = get_db_connection()
    
    # Use session filters
    filters = session.get('scan_filters', {})
    business_unit = filters.get('business_unit', '')
    scan_tool = filters.get('scan_tool', '')
    scan_type = filters.get('scan_type', '')
    most_recent_only = filters.get('most_recent_only', False)
    
    # Build query with same filters as the display page
    if most_recent_only:
        query = '''
            SELECT s.*, a.BusinessUnit, a.Rapid7App, a.CheckmarxProduct, a.MendProduct, a.MendProject
            FROM Scans s
            JOIN Artifacts a ON s.ArtifactID = a.ID
            WHERE s.ID IN (
                SELECT s2.ID FROM Scans s2
                WHERE s2.ArtifactID = s.ArtifactID 
                  AND s2.ScanTool = s.ScanTool
                  AND s2.ScanType = s.ScanType
                  AND s2.ScanDateTime = (
                    SELECT MAX(s3.ScanDateTime) 
                    FROM Scans s3
                    WHERE s3.ArtifactID = s2.ArtifactID
                      AND s3.ScanTool = s2.ScanTool
                      AND s3.ScanType = s2.ScanType
                  )
            )
        '''
    else:
        query = '''
            SELECT s.*, a.BusinessUnit, a.Rapid7App, a.CheckmarxProduct, a.MendProduct, a.MendProject
            FROM Scans s
            JOIN Artifacts a ON s.ArtifactID = a.ID
            WHERE 1=1
        '''
    params = []
    
    if business_unit:
        query += ' AND a.BusinessUnit = ?'
        params.append(business_unit)
    
    if scan_tool:
        query += ' AND s.ScanTool LIKE ?'
        params.append(f'%{scan_tool}%')
    
    if scan_type:
        query += ' AND s.ScanType LIKE ?'
        params.append(f'%{scan_type}%')
    
    query += ' ORDER BY s.ScanDateTime DESC'
    
    scans = conn.execute(query, params).fetchall()
    conn.close()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    
    # Generate sheet name based on filters
    sheet_parts = []
    if business_unit:
        sheet_parts.append(business_unit)
    else:
        sheet_parts.append('AllBUs')
    
    if scan_tool:
        sheet_parts.append(f'Tool_{scan_tool}')
    if scan_type:
        sheet_parts.append(f'Type_{scan_type}')
    if most_recent_only:
        sheet_parts.append('MostRecent')
    
    sheet_name = '_'.join(sheet_parts)[:31]  # Excel sheet name limit
    ws.title = sheet_name
    
    # Add headers
    headers = ['ID', 'Business Unit', 'Rapid7 App', 'Checkmarx Product', 'Mend Product', 'Mend Project',
               'Scan Tool', 'Scan Type', 'Scan DateTime', 'Repeat Count',
               'Critical', 'High', 'Medium', 'Critical NP', 'High NP', 'Medium NP']
    ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.font = Font(bold=True)
    
    # Add data
    for scan in scans:
        ws.append([
            scan['ID'],
            scan['BusinessUnit'],
            scan['Rapid7App'] or '',
            scan['CheckmarxProduct'] or '',
            scan['MendProduct'] or '',
            scan['MendProject'] or '',
            scan['ScanTool'],
            scan['ScanType'],
            scan['ScanDateTime'],
            scan['ScanRepeatCount'],
            scan['Critical'],
            scan['High'],
            scan['Medium'],
            scan['CriticalNP'],
            scan['HighNP'],
            scan['MediumNP']
        ])
    
    # Freeze top row
    ws.freeze_panes = 'A2'
    
    # Add autofilter
    ws.auto_filter.ref = ws.dimensions
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'Scans_{sheet_name}_{timestamp}.xlsx'
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/scans/new', methods=['GET', 'POST'])
def new_scan():
    conn = get_db_connection()
    
    if request.method == 'POST':
        try:
            conn.execute('''
                INSERT INTO Scans (
                    ArtifactID, ScanTool, ScanType, ScanDateTime, ScanRepeatCount,
                    Critical, High, Medium, CriticalNP, HighNP, MediumNP
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                int(request.form['artifact_id']),
                request.form['scan_tool'],
                request.form['scan_type'],
                request.form['scan_datetime'],
                int(request.form.get('scan_repeat_count', 1)),
                int(request.form.get('critical', 0)),
                int(request.form.get('high', 0)),
                int(request.form.get('medium', 0)),
                int(request.form.get('critical_np', 0)),
                int(request.form.get('high_np', 0)),
                int(request.form.get('medium_np', 0))
            ))
            conn.commit()
            conn.close()
            flash('Scan created successfully!', 'success')
            return redirect(url_for('scans'))
        except sqlite3.IntegrityError as e:
            flash(f'Database error: {str(e)}', 'error')
        except Exception as e:
            flash(f'Error: {str(e)}', 'error')
    
    artifacts = conn.execute('SELECT ID, BusinessUnit, Rapid7App, CheckmarxProduct, MendProduct, MendProject FROM Artifacts WHERE Deleted = 0 ORDER BY BusinessUnit').fetchall()
    conn.close()
    return render_template('scan_form.html', scan=None, artifacts=artifacts)

@app.route('/scans/<int:id>/edit', methods=['GET', 'POST'])
def edit_scan(id):
    conn = get_db_connection()
    
    if request.method == 'POST':
        try:
            conn.execute('''
                UPDATE Scans SET
                    ArtifactID = ?, ScanTool = ?, ScanType = ?, ScanDateTime = ?, ScanRepeatCount = ?,
                    Critical = ?, High = ?, Medium = ?, CriticalNP = ?, HighNP = ?, MediumNP = ?
                WHERE ID = ?
            ''', (
                int(request.form['artifact_id']),
                request.form['scan_tool'],
                request.form['scan_type'],
                request.form['scan_datetime'],
                int(request.form.get('scan_repeat_count', 1)),
                int(request.form.get('critical', 0)),
                int(request.form.get('high', 0)),
                int(request.form.get('medium', 0)),
                int(request.form.get('critical_np', 0)),
                int(request.form.get('high_np', 0)),
                int(request.form.get('medium_np', 0)),
                id
            ))
            conn.commit()
            flash('Scan updated successfully!', 'success')
            return redirect(url_for('scans'))
        except sqlite3.IntegrityError as e:
            flash(f'Database error: {str(e)}', 'error')
        except Exception as e:
            flash(f'Error: {str(e)}', 'error')
    
    scan = conn.execute('SELECT * FROM Scans WHERE ID = ?', (id,)).fetchone()
    artifacts = conn.execute('SELECT ID, BusinessUnit, Rapid7App, CheckmarxProduct, MendProduct, MendProject FROM Artifacts WHERE Deleted = 0 ORDER BY BusinessUnit').fetchall()
    conn.close()
    
    if scan is None:
        flash('Scan not found!', 'error')
        return redirect(url_for('scans'))
    
    return render_template('scan_form.html', scan=scan, artifacts=artifacts)

@app.route('/scans/<int:id>/delete', methods=['POST'])
def delete_scan(id):
    conn = get_db_connection()
    conn.execute('DELETE FROM Scans WHERE ID = ?', (id,))
    conn.commit()
    conn.close()
    flash('Scan deleted!', 'success')
    return redirect(url_for('scans'))

@app.route('/artifacts/<int:id>/scans')
def artifact_scans(id):
    conn = get_db_connection()
    artifact = conn.execute('SELECT * FROM Artifacts WHERE ID = ?', (id,)).fetchone()
    
    if artifact is None:
        flash('Artifact not found!', 'error')
        conn.close()
        return redirect(url_for('artifacts'))
    
    scans = conn.execute('''
        SELECT * FROM Scans WHERE ArtifactID = ? ORDER BY ScanDateTime DESC
    ''', (id,)).fetchall()
    conn.close()
    return render_template('artifact_scans.html', artifact=artifact, scans=scans)

if __name__ == '__main__':
    app.run(debug=True, port=5000)
